from __future__ import annotations

import asyncio
import csv
import logging
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from aiogram import Bot
from aiogram.types import BufferedInputFile, InlineKeyboardButton, InlineKeyboardMarkup

from bot.config import get_settings
from bot.db.base import get_session_maker
from bot.db.models import Client, Contract, ContractStatus, Payment, PaymentStatus, ReminderRepeat, ReminderStatus, User, UserRole

log = logging.getLogger(__name__)


_MARKER_PREFIX = "PAYMENTS_SWEEP:"


@dataclass(frozen=True)
class PaymentRow:
    contract_id: int
    contract_number: str
    contract_company: str
    contract_kind: str
    currency: str
    due_date: date
    amount_minor: int
    client_id: int
    client_name: str
    client_phone: str | None

    @property
    def amount(self) -> float:
        return self.amount_minor / 100.0


@dataclass(frozen=True)
class ContractEndRow:
    contract_id: int
    contract_number: str
    contract_company: str
    contract_kind: str
    currency: str
    end_date: date
    client_id: int
    client_name: str
    client_phone: str | None
    pending_payments_count: int


def _fmt_money(amount: float, currency: str) -> str:
    # Keep consistent float formatting used across UI.
    return f"{amount:.2f} {currency}"


def _payments_totals_by_currency(payments: list[PaymentRow]) -> str:
    totals_minor: dict[str, int] = defaultdict(int)
    for p in payments:
        totals_minor[p.currency] += p.amount_minor
    parts: list[str] = []
    for currency in sorted(totals_minor.keys()):
        parts.append(f"{totals_minor[currency] / 100.0:.2f} {currency}")
    return ", ".join(parts)


def _payments_to_csv_bytes(
    payments: list[PaymentRow],
    *,
    days_ahead: int,
    due_date: date,
) -> tuple[str, bytes]:
    """
    Export payments into CSV for Telegram document sending.
    Uses UTF-8 BOM to make Excel happy.
    """
    # csv module writes to str; encode to bytes for BufferedInputFile.
    from io import StringIO

    out = StringIO()
    writer = csv.writer(out, delimiter=",", quotechar='"')
    writer.writerow(["дата платежа", "клиент", "номер договора", "сумма", "валюта", "телефон"])
    for p in sorted(payments, key=lambda x: (x.due_date, x.client_name, x.contract_number)):
        phone_cell = ""
        if p.client_phone:
            # Make it clickable when opened in Excel.
            phone_cell = f'=HYPERLINK("tel:{p.client_phone}", "{p.client_phone}")'
        writer.writerow(
            [
                p.due_date.isoformat(),
                p.client_name,
                p.contract_number,
                f"{p.amount_minor/100.0:.2f}",
                p.currency,
                phone_cell,
            ]
        )
    filename = f"payments_due_{days_ahead}d_{due_date:%Y-%m-%d}.csv"
    return filename, out.getvalue().encode("utf-8-sig")


def _payments_workbook_xlsx_bytes(payments: list[PaymentRow], filename: str) -> tuple[str, bytes]:
    """Build .xlsx (Excel) with clickable phone links; returns (filename, raw bytes)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Взносы"

    headers = ["дата платежа", "клиент", "номер договора", "сумма", "валюта", "телефон"]
    ws.append(headers)

    sorted_payments = sorted(payments, key=lambda x: (x.due_date, x.client_name, x.contract_number))
    for p in sorted_payments:
        amount = p.amount_minor / 100.0
        ws.append([p.due_date.isoformat(), p.client_name, p.contract_number, amount, p.currency, p.client_phone or ""])

    phone_col_letter = "F"
    link_font = Font(color="0000FF", underline="single")
    for row_idx in range(2, 2 + len(sorted_payments)):
        cell = ws[f"{phone_col_letter}{row_idx}"]
        phone = str(cell.value or "").strip()
        if phone:
            cell.hyperlink = f"tel:{phone}"
            cell.font = link_font

    width_caps = {"A": 12, "B": 32, "C": 22, "D": 14, "E": 10, "F": 18}
    for col_letter, cap in width_caps.items():
        max_len = 0
        for cell in ws[col_letter]:
            val = cell.value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, cap)

    for row_idx in range(2, 2 + len(sorted_payments)):
        ws[f"B{row_idx}"].alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    return filename, bio.getvalue()


def _payments_period_to_xlsx_bytes(
    payments: list[PaymentRow],
    *,
    date_from: date,
    date_to: date,
) -> tuple[str, bytes]:
    """Inclusive date range [date_from, date_to]."""
    filename = f"payments_period_{date_from:%Y-%m-%d}_{date_to:%Y-%m-%d}.xlsx"
    return _payments_workbook_xlsx_bytes(payments, filename)


def _contract_ends_period_to_xlsx_bytes(
    contracts: list[ContractEndRow],
    *,
    date_from: date,
    date_to: date,
) -> tuple[str, bytes]:
    """Contracts ending in [date_from, date_to], exported as xlsx."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Договоры"

    headers = [
        "дата окончания",
        "клиент",
        "номер договора",
        "компания",
        "вид",
        "валюта",
        "ожидающих взносов",
        "телефон",
    ]
    ws.append(headers)

    rows = sorted(contracts, key=lambda x: (x.end_date, x.client_name, x.contract_number))
    for c in rows:
        ws.append(
            [
                c.end_date.isoformat(),
                c.client_name,
                c.contract_number,
                c.contract_company,
                c.contract_kind,
                c.currency,
                c.pending_payments_count,
                c.client_phone or "",
            ]
        )

    link_font = Font(color="0000FF", underline="single")
    for row_idx in range(2, 2 + len(rows)):
        cell = ws[f"H{row_idx}"]
        phone = str(cell.value or "").strip()
        if phone:
            cell.hyperlink = f"tel:{phone}"
            cell.font = link_font

    width_caps = {"A": 14, "B": 32, "C": 24, "D": 20, "E": 20, "F": 10, "G": 18, "H": 18}
    for col_letter, cap in width_caps.items():
        max_len = 0
        for cell in ws[col_letter]:
            val = cell.value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, cap)

    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    filename = f"contracts_ending_{date_from:%Y-%m-%d}_{date_to:%Y-%m-%d}.xlsx"
    return filename, bio.getvalue()


async def get_pending_payments_due_between(agent_tg_id: int, start: date, end: date) -> list[PaymentRow]:
    """Pending payments with due_date in [start, end] inclusive (agent's clients only)."""
    from sqlalchemy import select

    if end < start:
        start, end = end, start

    async with get_session_maker()() as session:
        res = await session.execute(
            select(Payment, Contract, Client)
            .join(Contract, Contract.id == Payment.contract_id)
            .join(Client, Client.id == Contract.client_id)
            .join(User, User.id == Client.agent_user_id)
            .where(
                User.tg_id == agent_tg_id,
                Payment.status == PaymentStatus.pending,
                Contract.status == ContractStatus.active,
                Payment.due_date >= start,
                Payment.due_date <= end,
            )
            .order_by(Payment.due_date.asc(), Client.full_name.asc(), Contract.contract_number.asc())
        )
        rows: list[PaymentRow] = []
        for payment, contract, client in res.all():
            rows.append(
                PaymentRow(
                    contract_id=contract.id,
                    contract_number=contract.contract_number,
                    contract_company=contract.company,
                    contract_kind=contract.contract_kind,
                    currency=contract.currency,
                    due_date=payment.due_date,
                    amount_minor=payment.amount_minor,
                    client_id=client.id,
                    client_name=client.full_name,
                    client_phone=client.phone,
                )
            )
        return rows


async def get_contracts_ending_between(agent_tg_id: int, start: date, end: date) -> list[ContractEndRow]:
    """Contracts with end_date in [start, end] for current agent."""
    from sqlalchemy import func, select

    if end < start:
        start, end = end, start

    async with get_session_maker()() as session:
        res = await session.execute(
            select(Contract, Client, func.count(Payment.id))
            .join(Client, Client.id == Contract.client_id)
            .join(User, User.id == Client.agent_user_id)
            .outerjoin(
                Payment,
                (Payment.contract_id == Contract.id) & (Payment.status == PaymentStatus.pending),
            )
            .where(
                User.tg_id == agent_tg_id,
                Contract.status == ContractStatus.active,
                Contract.end_date >= start,
                Contract.end_date <= end,
            )
            .group_by(Contract.id, Client.id)
            .order_by(Contract.end_date.asc(), Client.full_name.asc(), Contract.contract_number.asc())
        )
        rows: list[ContractEndRow] = []
        for contract, client, pending_count in res.all():
            rows.append(
                ContractEndRow(
                    contract_id=contract.id,
                    contract_number=contract.contract_number,
                    contract_company=contract.company,
                    contract_kind=contract.contract_kind,
                    currency=contract.currency,
                    end_date=contract.end_date,
                    client_id=client.id,
                    client_name=client.full_name,
                    client_phone=client.phone,
                    pending_payments_count=int(pending_count or 0),
                )
            )
        return rows


def _format_payments_due_report(
    payments: list[PaymentRow],
    *,
    days_ahead: int,
    due_date: date,
    offset: int = 0,
    page_size: int = 30,
    max_chars: int = 3500,
) -> str:
    header = f"⏳ Взносы через {days_ahead} день(ей) — {due_date:%d.%m.%Y}"
    total = len(payments)
    if total == 0:
        return f"{header}\n\n— нет ожидающих платежей"

    payments_page = payments[offset : offset + page_size]
    shown = 0

    lines: list[str] = [header, ""]
    current_len = len("\n".join(lines))
    for p in payments_page:
        amount = _fmt_money(p.amount, p.currency)
        line = f"• {p.client_name} → {p.contract_number}: {amount}"
        # Telegram hard limit is 4096 chars; keep a safe margin.
        if current_len + len(line) + 1 > max_chars:
            break
        lines.append(line)
        current_len += len(line) + 1
        shown += 1

    # Footer helps user understand which part they see.
    start_no = offset + 1
    end_no = offset + shown
    lines.append("")
    lines.append(f"Показано {start_no}-{end_no} из {total}.")

    remaining = total - end_no
    if remaining > 0:
        lines.append(f"... и ещё {remaining} платежей")

    return "\n".join(lines)


async def _marker_exists(session, agent_user_id: int, marker_text: str) -> bool:
    from sqlalchemy import select
    from bot.db.models import Reminder

    res = await session.execute(
        select(Reminder.id).where(Reminder.agent_user_id == agent_user_id, Reminder.text == marker_text).limit(1)
    )
    return res.scalar_one_or_none() is not None


async def _create_sent_marker(session, agent_user_id: int, marker_text: str) -> None:
    from sqlalchemy import select
    from bot.db.models import Reminder

    now = datetime.now(timezone.utc)
    # Note: reminders_worker sends only pending reminders, so sent markers won't be delivered to user.
    r = Reminder(
        agent_user_id=agent_user_id,
        text=marker_text,
        remind_at=now,
        status=ReminderStatus.sent,
        repeat=ReminderRepeat.none,
        sent_at=now,
    )
    session.add(r)
    # Caller should commit.


async def _fetch_agents(session) -> list[User]:
    from sqlalchemy import select

    res = await session.execute(select(User).where(User.role == UserRole.agent))
    return list(res.scalars().all())


async def _fetch_payments_due_up_to(session, due_end: date) -> list[tuple[Payment, Contract, Client, User]]:
    """
    Fetch pending payments for all agents where due_date <= due_end.
    (Overdue and upcoming are both included; caller groups by date.)
    """
    from sqlalchemy import select

    res = await session.execute(
        select(Payment, Contract, Client, User)
        .join(Contract, Contract.id == Payment.contract_id)
        .join(Client, Client.id == Contract.client_id)
        .join(User, User.id == Client.agent_user_id)
        .where(User.role == UserRole.agent, Payment.status == PaymentStatus.pending, Payment.due_date <= due_end)
        .order_by(Payment.due_date.asc())
    )
    return list(res.all())


async def _fetch_contract_ends_for_deltas(session, deltas: list[int], today: date) -> list[tuple[Contract, User, int]]:
    """
    Returns rows: (contract, agent_user, pending_payments_count) for contracts
    ending exactly at today + delta and having at least one pending payment.
    """
    from sqlalchemy import select, func

    target_dates = [today + timedelta(days=d) for d in deltas]

    res = await session.execute(
        select(Contract, User, func.count(Payment.id))
        .join(Client, Client.id == Contract.client_id)
        .join(User, User.id == Client.agent_user_id)
        .join(Payment, Payment.contract_id == Contract.id)
        .where(
            User.role == UserRole.agent,
            Payment.status == PaymentStatus.pending,
            Contract.end_date.in_(target_dates),
        )
        .group_by(Contract.id, User.id)
        .order_by(Contract.end_date.asc())
    )
    return [(row[0], row[1], int(row[2])) for row in res.all()]


def _build_payment_lines(payments: list[PaymentRow], limit: int = 20) -> list[str]:
    # Sort by due_date (already sorted in SQL, but keep deterministic).
    payments = sorted(payments, key=lambda p: (p.due_date, p.contract_id))
    lines: list[str] = []
    for p in payments[:limit]:
        lines.append(f"• {p.due_date:%d.%m.%Y}: {p.contract_number} — {_fmt_money(p.amount, p.currency)}")
    if len(payments) > limit:
        lines.append(f"… и ещё {len(payments) - limit} платежей")
    return lines


async def send_payment_reminder_sweep(bot: Bot) -> None:
    settings = get_settings()
    try:
        tz = ZoneInfo(settings.timezone)
    except ZoneInfoNotFoundError:
        log.warning("Timezone %r not found, falling back to UTC. Install tzdata on Windows.", settings.timezone)
        tz = ZoneInfo("UTC")

    local_now = datetime.now(tz)
    today = local_now.date()

    # Marker is per local day and per agent.
    marker_text = f"{_MARKER_PREFIX}{today.isoformat()}"

    # Fetch payments up to end of month window for lists.
    due_end = today + timedelta(days=29)

    async with get_session_maker()() as session:
        agents = await _fetch_agents(session)
        if not agents:
            return

        payments_all = await _fetch_payments_due_up_to(session, due_end=due_end)
        contracts_ends = await _fetch_contract_ends_for_deltas(session, deltas=[30, 14, 7], today=today)

        # Group payments by agent tg_id.
        payments_by_agent: dict[int, list[PaymentRow]] = defaultdict(list)
        for payment, contract, client, agent_user in payments_all:
            payments_by_agent[agent_user.tg_id].append(
                PaymentRow(
                    contract_id=contract.id,
                    contract_number=contract.contract_number,
                    contract_company=contract.company,
                    contract_kind=contract.contract_kind,
                    currency=contract.currency,
                    due_date=payment.due_date,
                    amount_minor=payment.amount_minor,
                    client_id=client.id,
                    client_name=client.full_name,
                    client_phone=client.phone,
                )
            )

        # Group contract ends by agent tg_id.
        contract_ends_by_agent: dict[int, list[tuple[Contract, int]]] = defaultdict(list)
        for contract, agent_user, pending_count in contracts_ends:
            contract_ends_by_agent[agent_user.tg_id].append((contract, pending_count))

        # Send per agent.
        for agent in agents:
            agent_tg_id = agent.tg_id
            agent_user_id = agent.id

            if await _marker_exists(session, agent_user_id, marker_text):
                continue

            # Get payments for this agent (might be empty).
            p_rows = payments_by_agent.get(agent_tg_id, [])

            # Split dates for +1/+3/+7 reports (inclusive window [today, today+N]) and other buckets.
            overdue = [p for p in p_rows if p.due_date < today]
            due_1 = [p for p in p_rows if today <= p.due_date <= today + timedelta(days=1)]
            due_3 = [p for p in p_rows if today <= p.due_date <= today + timedelta(days=3)]
            due_7 = [p for p in p_rows if today <= p.due_date <= today + timedelta(days=7)]

            week = [p for p in p_rows if today <= p.due_date <= today + timedelta(days=6)]
            month = [p for p in p_rows if today <= p.due_date <= today + timedelta(days=29)]

            # Contract end notifications.
            ends_rows = contract_ends_by_agent.get(agent_tg_id, [])
            ends_30 = [c for c, _cnt in ends_rows if c.end_date == today + timedelta(days=30)]
            ends_14 = [c for c, _cnt in ends_rows if c.end_date == today + timedelta(days=14)]
            ends_7 = [c for c, _cnt in ends_rows if c.end_date == today + timedelta(days=7)]

            # Build and send messages only if there is something to report.
            any_sent = False

            if overdue:
                any_sent = True
                lines = _build_payment_lines(overdue)
                text = "⛔ Просроченные платежи\n" + "\n".join(lines)
                await bot.send_message(agent_tg_id, text)

            # 1/3/7 daily button reports:
            # Requirement: 3 different buttons, each sends one report.
            # Also: every morning auto-send these three reports.
            due_1_count, due_3_count, due_7_count = len(due_1), len(due_3), len(due_7)
            any_sent = True

            end_1, end_3, end_7 = today + timedelta(days=1), today + timedelta(days=3), today + timedelta(days=7)
            summary_text = (
                f"⏰ Сводка по взносам\n"
                f"{today:%d.%m}—{end_1:%d.%m}: {due_1_count} взносов, {_payments_totals_by_currency(due_1)}\n"
                f"{today:%d.%m}—{end_3:%d.%m}: {due_3_count} взносов, {_payments_totals_by_currency(due_3)}\n"
                f"{today:%d.%m}—{end_7:%d.%m}: {due_7_count} взносов, {_payments_totals_by_currency(due_7)}"
            )
            kb = InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(text="+1 день", callback_data="payrep:1"),
                        InlineKeyboardButton(text="+3 дня", callback_data="payrep:3"),
                        InlineKeyboardButton(text="+7 дней", callback_data="payrep:7"),
                    ],
                    [
                        InlineKeyboardButton(text="30 дней", callback_data="payrep:range:30"),
                    ],
                ]
            )
            await bot.send_message(agent_tg_id, summary_text, reply_markup=kb)

            # Auto export 3 detailed reports as files for reliability.
            if due_1:
                fn, payload = _payments_period_to_xlsx_bytes(due_1, date_from=today, date_to=end_1)
                await bot.send_document(
                    agent_tg_id,
                    BufferedInputFile(payload, filename=fn),
                    caption=(
                        f"⏳ Период {today:%d.%m.%Y}—{end_1:%d.%m.%Y}. "
                        f"Всего: {due_1_count} взносов, {_payments_totals_by_currency(due_1)}"
                    ),
                )
            else:
                await bot.send_message(
                    agent_tg_id,
                    f"⏳ Период {today:%d.%m.%Y}—{end_1:%d.%m.%Y}: нет ожидающих взносов.",
                )

            if due_3:
                fn, payload = _payments_period_to_xlsx_bytes(due_3, date_from=today, date_to=end_3)
                await bot.send_document(
                    agent_tg_id,
                    BufferedInputFile(payload, filename=fn),
                    caption=(
                        f"⏳ Период {today:%d.%m.%Y}—{end_3:%d.%m.%Y}. "
                        f"Всего: {due_3_count} взносов, {_payments_totals_by_currency(due_3)}"
                    ),
                )
            else:
                await bot.send_message(
                    agent_tg_id,
                    f"⏳ Период {today:%d.%m.%Y}—{end_3:%d.%m.%Y}: нет ожидающих взносов.",
                )

            if due_7:
                fn, payload = _payments_period_to_xlsx_bytes(due_7, date_from=today, date_to=end_7)
                await bot.send_document(
                    agent_tg_id,
                    BufferedInputFile(payload, filename=fn),
                    caption=(
                        f"⏳ Период {today:%d.%m.%Y}—{end_7:%d.%m.%Y}. "
                        f"Всего: {due_7_count} взносов, {_payments_totals_by_currency(due_7)}"
                    ),
                )
            else:
                await bot.send_message(
                    agent_tg_id,
                    f"⏳ Период {today:%d.%m.%Y}—{end_7:%d.%m.%Y}: нет ожидающих взносов.",
                )

            if ends_rows and (ends_30 or ends_14 or ends_7):
                any_sent = True
                text_parts = ["📅 До конца договора с pending платежами:"]
                if ends_30:
                    text_parts.append(f"• через 30 дней: {len(ends_30)} договор(ов)")
                if ends_14:
                    text_parts.append(f"• через 14 дней: {len(ends_14)} договор(ов)")
                if ends_7:
                    text_parts.append(f"• через 7 дней: {len(ends_7)} договор(ов)")
                await bot.send_message(agent_tg_id, "\n".join(text_parts))

            if week:
                any_sent = True
                end_week = today + timedelta(days=6)
                text = f"📋 Список платежей на неделю до {end_week:%d.%m.%Y}\n"
                text += "\n".join(_build_payment_lines(week, limit=15))
                await bot.send_message(agent_tg_id, text)

            if month:
                any_sent = True
                end_month = today + timedelta(days=29)
                text = f"📋 Список платежей на месяц до {end_month:%d.%m.%Y}\n"
                text += "\n".join(_build_payment_lines(month, limit=20))
                await bot.send_message(agent_tg_id, text)

            # Mark sweep done for this agent even if there was nothing.
            await _create_sent_marker(session, agent_user_id, marker_text)

        await session.commit()


async def payment_reminders_worker(bot: Bot, *, run_hour: int = 9, run_minute: int = 0) -> None:
    """
    Daily morning worker:
    - runs send_payment_reminder_sweep once per local day at `run_hour:run_minute`
    - marker avoids duplicates per day/agent.
    """
    settings = get_settings()
    try:
        tz = ZoneInfo(settings.timezone)
    except ZoneInfoNotFoundError:
        log.warning("Timezone %r not found, falling back to UTC. Install tzdata on Windows.", settings.timezone)
        tz = ZoneInfo("UTC")

    while True:
        try:
            now_local = datetime.now(tz)
            run_at = now_local.replace(hour=run_hour, minute=run_minute, second=0, microsecond=0)
            if run_at <= now_local:
                run_at = run_at + timedelta(days=1)

            sleep_seconds = (run_at - now_local).total_seconds()
            if sleep_seconds > 0:
                await asyncio.sleep(sleep_seconds)

            await send_payment_reminder_sweep(bot)
        except Exception as e:
            log.exception("Payment reminder worker error: %s", e)

