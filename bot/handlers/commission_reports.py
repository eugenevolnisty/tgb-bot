from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from aiogram import F, Router
from aiogram.exceptions import TelegramBadRequest
from aiogram.types import BufferedInputFile, CallbackQuery
from openpyxl import Workbook
from sqlalchemy import select

from bot.config import get_settings
from bot.db.base import get_session_maker
from bot.db.models import AgentCommission, Client, Contract, Payment, PaymentStatus, User

router = Router()


@dataclass(frozen=True)
class CommissionRow:
    paid_date_local: date
    client_name: str
    contract_number: str
    company: str
    contract_kind: str
    currency: str
    payment_amount_minor: int
    percent_bp: int
    commission_minor: int


def _commission_to_xlsx(rows: list[CommissionRow], date_from: date, date_to: date) -> tuple[str, bytes]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Детализация"
    ws.append(["Дата оплаты", "Клиент", "Договор", "Компания", "Вид", "Взнос", "Валюта", "Комиссия %", "Комиссия"])
    for r in rows:
        ws.append(
            [
                r.paid_date_local.isoformat(),
                r.client_name,
                r.contract_number,
                r.company,
                r.contract_kind,
                r.payment_amount_minor / 100.0,
                r.currency,
                r.percent_bp / 100.0,
                r.commission_minor / 100.0,
            ]
        )
    for col, width in {"A": 14, "B": 28, "C": 20, "D": 22, "E": 24, "F": 14, "G": 10, "H": 12, "I": 14}.items():
        ws.column_dimensions[col].width = width
    by_company = wb.create_sheet("По компаниям")
    by_company.append(["Компания", "Оплат", "Комиссия"])
    comp: dict[str, tuple[int, int]] = {}
    for r in rows:
        cnt, total = comp.get(r.company, (0, 0))
        comp[r.company] = (cnt + 1, total + r.commission_minor)
    for company in sorted(comp.keys()):
        cnt, total = comp[company]
        by_company.append([company, cnt, total / 100.0])
    by_kind = wb.create_sheet("По видам")
    by_kind.append(["Вид", "Оплат", "Комиссия"])
    kinds: dict[str, tuple[int, int]] = {}
    for r in rows:
        cnt, total = kinds.get(r.contract_kind, (0, 0))
        kinds[r.contract_kind] = (cnt + 1, total + r.commission_minor)
    for kind in sorted(kinds.keys()):
        cnt, total = kinds[kind]
        by_kind.append([kind, cnt, total / 100.0])
    fn = f"commissions_{date_from:%Y-%m-%d}_{date_to:%Y-%m-%d}.xlsx"
    bio = BytesIO()
    wb.save(bio)
    return fn, bio.getvalue()


async def get_commission_rows_for_period(agent_tg_id: int, date_from: date, date_to: date, tz: ZoneInfo) -> list[CommissionRow]:
    start_local = datetime.combine(date_from, time.min, tzinfo=tz)
    end_local = datetime.combine(date_to, time.max, tzinfo=tz)
    start_utc = start_local.astimezone(timezone.utc)
    end_utc = end_local.astimezone(timezone.utc)
    async with get_session_maker()() as session:
        q = (
            await session.execute(
                select(Payment, Contract, Client, AgentCommission)
                .join(Contract, Contract.id == Payment.contract_id)
                .join(Client, Client.id == Contract.client_id)
                .join(User, User.id == Client.agent_user_id)
                .join(
                    AgentCommission,
                    (AgentCommission.agent_user_id == User.id)
                    & (AgentCommission.company == Contract.company)
                    & (AgentCommission.contract_kind == Contract.contract_kind),
                )
                .where(
                    User.tg_id == agent_tg_id,
                    Payment.status == PaymentStatus.paid,
                    Payment.paid_at.is_not(None),
                    Payment.paid_at >= start_utc,
                    Payment.paid_at <= end_utc,
                )
                .order_by(Payment.paid_at.asc())
            )
        )
        out: list[CommissionRow] = []
        for p, c, cl, com in q.all():
            commission_minor = int(round(int(p.amount_minor) * int(com.percent_bp) / 10000.0))
            paid_date_local = p.paid_at.astimezone(tz).date() if p.paid_at else date_from
            out.append(
                CommissionRow(
                    paid_date_local=paid_date_local,
                    client_name=cl.full_name,
                    contract_number=c.contract_number,
                    company=c.company,
                    contract_kind=c.contract_kind,
                    currency=c.currency,
                    payment_amount_minor=int(p.amount_minor),
                    percent_bp=int(com.percent_bp),
                    commission_minor=commission_minor,
                )
            )
        return out


@router.callback_query(F.data.startswith("comrep:"))
async def commissions_report(callback: CallbackQuery) -> None:
    if callback.message is None:
        await callback.answer()
        return
    settings = get_settings()
    try:
        tz = ZoneInfo(settings.timezone)
    except ZoneInfoNotFoundError:
        tz = ZoneInfo("UTC")
    try:
        days = int(callback.data.split(":")[1])
    except Exception:
        await callback.answer("Некорректные данные", show_alert=True)
        return
    if days not in (1, 3, 7, 30):
        await callback.answer("Некорректные данные", show_alert=True)
        return
    today = datetime.now(tz).date()
    date_from = today - timedelta(days=days - 1)
    rows = await get_commission_rows_for_period(callback.from_user.id, date_from, today, tz)
    if not rows:
        await callback.message.answer(f"💼 Комиссии за период {date_from:%d.%m.%Y}—{today:%d.%m.%Y}: нет.")
        await callback.answer()
        return
    total_minor = sum(r.commission_minor for r in rows)
    totals_by_currency: dict[str, int] = {}
    for r in rows:
        totals_by_currency[r.currency] = totals_by_currency.get(r.currency, 0) + r.commission_minor
    parts = [f"{v/100.0:.2f} {k}" for k, v in sorted(totals_by_currency.items())]
    fn, payload = _commission_to_xlsx(rows, date_from, today)
    caption = (
        f"💼 Комиссии за период {date_from:%d.%m.%Y}—{today:%d.%m.%Y}. "
        f"Всего: {len(rows)} оплат, комиссия {total_minor/100.0:.2f} ({', '.join(parts)})"
    )
    try:
        await callback.message.answer_document(
            document=BufferedInputFile(payload, filename=fn),
            caption=caption,
        )
    except TelegramBadRequest:
        await callback.message.answer(f"Не удалось отправить файл. Всего оплат: {len(rows)}")
    await callback.answer()

